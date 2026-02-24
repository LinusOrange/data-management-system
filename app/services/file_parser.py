from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.models import SourceType


@dataclass
class ParsedPreviewRow:
    row_no: int
    name: str | None
    item_code: str | None
    qty: Decimal | None
    amount: Decimal | None
    order_no: str | None
    raw: dict


def _normalize_header(value: str | None) -> str:
    if not value:
        return ""
    return "".join(str(value).strip().replace("\n", "").split())


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return None


def _mapping_for_source(source_type: SourceType) -> dict[str, str]:
    if source_type == SourceType.inbound:
        return {
            "商品名称": "name",
            "型号": "item_code",
            "数量": "qty",
            "折前金额": "amount",
            "摘要": "order_no",
        }
    return {
        "销售订单": "order_no",
        "产品": "item_code",
        "产品名称": "name",
        "销售数量": "qty",
        "含税金额": "amount",
    }


def parse_file_to_preview_rows(file_path: Path, source_type: SourceType) -> list[ParsedPreviewRow]:
    ext = file_path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _parse_xlsx(file_path, source_type)
    if ext == ".xls":
        return _parse_xls(file_path, source_type)
    if ext == ".csv":
        return _parse_csv(file_path, source_type)
    raise ValueError(f"unsupported file extension: {ext}")





def _has_required_inbound_fields(row_dict: dict) -> bool:
    required_keys = ["name", "item_code", "qty", "amount", "order_no"]
    for key in required_keys:
        value = row_dict.get(key)
        if value is None:
            return False
        if isinstance(value, str) and value.strip() == "":
            return False
    return True

def _parse_xls(file_path: Path, source_type: SourceType) -> list[ParsedPreviewRow]:
    try:
        import xlrd
    except ImportError as exc:  # noqa: PERF203
        raise ValueError("xls parsing requires 'xlrd' dependency") from exc

    workbook = xlrd.open_workbook(file_path.as_posix())
    sheet = workbook.sheet_by_index(0)

    if source_type == SourceType.inbound:
        return _parse_inbound_xls_fixed_header(sheet)

    return _parse_xls_by_header_scan(sheet, source_type)


def _parse_inbound_xls_fixed_header(sheet) -> list[ParsedPreviewRow]:
    """Inbound .xls template: row 17 header and E/H/M/P/W values."""
    col_idx = {
        "name": 4,
        "item_code": 7,
        "qty": 12,
        "amount": 15,
        "order_no": 22,
    }

    expected_headers = {
        "name": "商品名称",
        "item_code": "型号",
        "qty": "数量",
        "amount": "金额",
        "order_no": "摘要",
    }

    header_row = 16
    header_errors: list[str] = []
    for field, idx in col_idx.items():
        raw_header = sheet.cell_value(header_row, idx) if idx < sheet.ncols else None
        normalized = _normalize_header(raw_header)
        expected = expected_headers[field]
        if expected not in normalized:
            column_label = chr(ord("A") + idx)
            header_errors.append(
                f"{column_label}17 expected contains '{expected}', actual='{normalized or 'EMPTY'}'"
            )

    if header_errors:
        raise ValueError("inbound header validation failed: " + "; ".join(header_errors))

    data_rows: list[ParsedPreviewRow] = []
    for row_idx in range(17, sheet.nrows):
        row_dict = {
            "name": sheet.cell_value(row_idx, col_idx["name"]) if col_idx["name"] < sheet.ncols else None,
            "item_code": sheet.cell_value(row_idx, col_idx["item_code"]) if col_idx["item_code"] < sheet.ncols else None,
            "qty": sheet.cell_value(row_idx, col_idx["qty"]) if col_idx["qty"] < sheet.ncols else None,
            "amount": sheet.cell_value(row_idx, col_idx["amount"]) if col_idx["amount"] < sheet.ncols else None,
            "order_no": sheet.cell_value(row_idx, col_idx["order_no"]) if col_idx["order_no"] < sheet.ncols else None,
            "raw_excel_row": row_idx + 1,
        }
        if not _has_required_inbound_fields(row_dict):
            continue
        preview = _build_preview_row_fixed(row_idx + 1, row_dict)
        if preview:
            data_rows.append(preview)

    return data_rows


def _parse_xls_by_header_scan(sheet, source_type: SourceType) -> list[ParsedPreviewRow]:
    mapping = _mapping_for_source(source_type)
    headers: list[str] = []
    header_found = False
    data_rows: list[ParsedPreviewRow] = []

    for row_idx in range(sheet.nrows):
        row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
        values = [str(v).strip() if v is not None else "" for v in row]
        norm_values = [_normalize_header(v) for v in values]

        if not header_found:
            if all(col in norm_values for col in mapping.keys()):
                headers = norm_values
                header_found = True
            continue

        if not any(v not in {"", None} for v in values):
            continue

        row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        preview = _build_preview_row(row_idx + 1, row_dict, mapping)
        if preview:
            data_rows.append(preview)

    return data_rows

def _parse_xlsx(file_path: Path, source_type: SourceType) -> list[ParsedPreviewRow]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active

    if source_type == SourceType.inbound:
        rows = _parse_inbound_xlsx_fixed_header(sheet)
        workbook.close()
        return rows

    rows = _parse_xlsx_by_header_scan(sheet, source_type)
    workbook.close()
    return rows


def _parse_inbound_xlsx_fixed_header(sheet) -> list[ParsedPreviewRow]:
    """Inbound template is fixed: row 17 is header and E/H/M/P/W are used for values."""
    # 1-based excel column indexes
    col_idx = {
        "name": 5,       # E 列: 商品名称 -> 名称
        "item_code": 8,  # H 列: 型号 -> 货号
        "qty": 13,       # M 列: 数量
        "amount": 16,    # P 列: 金额
        "order_no": 23,  # W 列: 摘要 -> 订单号
    }

    expected_headers = {
        "name": "商品名称",
        "item_code": "型号",
        "qty": "数量",
        "amount": "金额",
        "order_no": "摘要",
    }

    header_errors: list[str] = []
    for field, column in col_idx.items():
        raw_header = sheet.cell(row=17, column=column).value
        normalized = _normalize_header(raw_header)
        expected = expected_headers[field]
        if expected not in normalized:
            column_label = chr(ord("A") + column - 1)
            header_errors.append(
                f"{column_label}17 expected contains '{expected}', actual='{normalized or 'EMPTY'}'"
            )

    if header_errors:
        raise ValueError("inbound header validation failed: " + "; ".join(header_errors))

    data_rows: list[ParsedPreviewRow] = []
    for excel_row_no in range(18, sheet.max_row + 1):
        row_dict = {
            "name": sheet.cell(row=excel_row_no, column=col_idx["name"]).value,
            "item_code": sheet.cell(row=excel_row_no, column=col_idx["item_code"]).value,
            "qty": sheet.cell(row=excel_row_no, column=col_idx["qty"]).value,
            "amount": sheet.cell(row=excel_row_no, column=col_idx["amount"]).value,
            "order_no": sheet.cell(row=excel_row_no, column=col_idx["order_no"]).value,
            "raw_excel_row": excel_row_no,
        }
        if not _has_required_inbound_fields(row_dict):
            continue
        preview = _build_preview_row_fixed(excel_row_no, row_dict)
        if preview:
            data_rows.append(preview)

    return data_rows


def _build_preview_row_fixed(excel_row_no: int, row_dict: dict) -> ParsedPreviewRow | None:
    if all(row_dict.get(k) in (None, "") for k in ["name", "item_code", "qty", "amount", "order_no"]):
        return None

    return ParsedPreviewRow(
        row_no=excel_row_no,
        name=str(row_dict["name"]).strip() if row_dict.get("name") not in (None, "") else None,
        item_code=str(row_dict["item_code"]).strip() if row_dict.get("item_code") not in (None, "") else None,
        qty=_to_decimal(row_dict.get("qty")),
        amount=_to_decimal(row_dict.get("amount")),
        order_no=str(row_dict["order_no"]).strip() if row_dict.get("order_no") not in (None, "") else None,
        raw={k: (str(v) if v is not None else None) for k, v in row_dict.items()},
    )


def _parse_xlsx_by_header_scan(sheet, source_type: SourceType) -> list[ParsedPreviewRow]:
    mapping = _mapping_for_source(source_type)
    headers: list[str] = []
    header_found = False
    data_rows: list[ParsedPreviewRow] = []

    for excel_row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [str(v).strip() if v is not None else "" for v in row]
        norm_values = [_normalize_header(v) for v in values]

        if not header_found:
            if all(col in norm_values for col in mapping.keys()):
                headers = norm_values
                header_found = True
            continue

        if not any(v not in {"", None} for v in values):
            continue

        row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        preview = _build_preview_row(excel_row_no, row_dict, mapping)
        if preview:
            data_rows.append(preview)

    return data_rows


def _parse_csv(file_path: Path, source_type: SourceType) -> list[ParsedPreviewRow]:
    mapping = _mapping_for_source(source_type)
    rows: list[ParsedPreviewRow] = []

    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        normalized_headers = {_normalize_header(k): k for k in (reader.fieldnames or [])}
        if not all(header in normalized_headers for header in mapping):
            missing = [h for h in mapping if h not in normalized_headers]
            raise ValueError(f"missing headers in csv: {missing}")

        for idx, row in enumerate(reader, start=2):
            row_dict = {_normalize_header(k): v for k, v in row.items()}
            preview = _build_preview_row(idx, row_dict, mapping)
            if preview:
                rows.append(preview)

    return rows


def _build_preview_row(excel_row_no: int, row_dict: dict, mapping: dict[str, str]) -> ParsedPreviewRow | None:
    mapped: dict[str, object] = {"name": None, "item_code": None, "qty": None, "amount": None, "order_no": None}
    for source_col, target_col in mapping.items():
        mapped[target_col] = row_dict.get(source_col)

    if all(mapped[k] in (None, "") for k in ["name", "item_code", "qty", "amount", "order_no"]):
        return None

    return ParsedPreviewRow(
        row_no=excel_row_no,
        name=str(mapped["name"]).strip() if mapped["name"] not in (None, "") else None,
        item_code=str(mapped["item_code"]).strip() if mapped["item_code"] not in (None, "") else None,
        qty=_to_decimal(mapped["qty"]),
        amount=_to_decimal(mapped["amount"]),
        order_no=str(mapped["order_no"]).strip() if mapped["order_no"] not in (None, "") else None,
        raw={k: (str(v) if v is not None else None) for k, v in row_dict.items()},
    )
